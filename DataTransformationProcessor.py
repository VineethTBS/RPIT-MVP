import boto3
import formulas
import pandas as pd
from openpyxl import load_workbook
import string
import numpy as C
import collections
import re
import datetime
import dill
from io import BytesIO
class DealProcessor: 
    def __init__(self,filedata,poolId,dealId):# This function is called when the Lambda function is triggered
        self.fileData=filedata # This is the file data
        self.poolId=poolId # This is the pool id
        self.dealId=dealId


    def col2num(self,col): # This function converts the column name to the column number in excel
        num = 0
        
        for c in col:  # This loop is used to convert the column name to the column number in excel
            if c in string.ascii_letters: # This condition is used to check if the character is a letter
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1 # This is the formula to convert the column name to the column number in excel
        return num # This returns the column number in excel 


    # func = formulas.Parser().ast('=Data!A2+1')[1].compile()
    # dillData = dill.dumps(func)
    # print(type(dillData))
    # pickledData = pickle.dumps(formulas.Parser().ast('=Data!A2+1')[1].compile())

    # dyn_resource = boto3.resource('dynamodb')
    # table = dyn_resource.Table("Function")
    # client = boto3.client('dynamodb', aws_access_key_id='AKIARYPIHLWH2CRORNK5',
    #     aws_secret_access_key='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',
    #     )
    def createFunctionObject(self,formula): # This function is used to create the function object from the formula
        isConstant = False # This is a boolean variable to check if the formula is a constant
        vLookupHash = None # This is the hash value of the vlookup function
        if len(str(formula))==0 or str(formula)[0]!='=': # This condition is used to check if the formula is a constant
            isConstant = True # This is a boolean variable to check if the formula is a constant 
            ConstantVal = formula # This is the constant value of the formula
            return {"isConstant": isConstant,"ConstantVal": ConstantVal}    # This returns the constant value of the formula 
        func = formulas.Parser().ast(formula)[1].compile() # This is the function object of the formula  
        
        parameters = [] # This is the list of parameters of the function object
        for parameter in func.inputs: # This loop is used to get the parameters of the function object 
            isSelfRef = False # This is a boolean variable to check if the parameter is a self reference 
            isVlookup = False # This is a boolean variable to check if the parameter is a vlookup function
            if 'VLOOKUP' in  str(parameter): # This condition is used to check if the parameter is a vlookup function 
                isVlookup = True
                start,end = str(parameter).split('!')[1].split(":") # This is the start and end of the vlookup function 
         
                vLookupHash = dill.dumps(  self.getVlookup(start,end)) # This is the hash value of the vlookup function 

                ##to do: vlook up logic
            elif 'Data!'.lower() in str(parameter).lower(): # This condition is used to check if the parameter is a self reference 
                parameterIndex = self.col2num(str(parameter).split('!')[1]) # This is the column number of the parameter
            else:
                isSelfRef = True # This is a boolean variable to check if the parameter is a self reference
                parameterIndex = self.col2num(str(parameter)) # This is the column number of the parameter
            parameters.append({"parameterIndex":parameterIndex,"vlookUpHash":vLookupHash,"isSelfRef":isSelfRef,"isVlookup":isVlookup}) # This is the parameter of the function object
        return {"parameter": parameters,"isConstant":isConstant}


    def getDynamo(self): # This function is used to get the dynamo resource
        return boto3.resource( # This returns the dynamo resource
        'dynamodb',
       #  aws_access_key_id='ASIA4C44A42VL4KOE7WH',
       #  aws_secret_access_key='qWo/Qi+HsKT+5TKDLGyfItDjDkYlI1uOHVP392cc', 
       #  aws_session_token = "FwoGZXIvYXdzEC8aDEcv2SBYJ6bZu+VdzCLVAWlbX4PSDzh4pYCIE9XS2BQNdE8FeknOE0+y4aXXB/WSaoRnawUkWs0I0i1ZjtTifqOfHQ+/JWEmqT5hXbvXfoDWHaozOn85yWFlxC4FGXQ0lAhLhaG5toDVfYYtvqinJApGtDMnsR0HGh0LuxQs+YZpAeDNObPLiJvMa5efdXlG2wfr7neu0cFSCXoHLv3sEdXRRjXLZ7wC+yu9eDU+WRfWn+e4cj0E5q/YvQzypOVUjNpMqUu7lg4ViEvcmjZaCCloK1F8uPk4G5R5Ur33D87ZpkyooyiY3KieBjIrpCL6FJFs3SddMiJiHyASLz9MziuyC2cXjnPxSZ/6RYfhv3NFWZGO9LuGqg==",
         region_name="us-east-1")
        

    def PrimaryDealData(self): # This function is used to process the primary deal data
        wb = load_workbook(BytesIO( self.fileData))# This is the workbook of the primary deal data 
        sheet_names = wb.get_sheet_names() # This is the sheet names of the primary deal data
        inputId = str(datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')) # This is the input id of the primary deal data
       
        sheet_ranges = wb["Data"] # This is the data sheet of the primary deal data 
        dfInput = pd.DataFrame(sheet_ranges.values)# if you want to include the header row
        new_header = dfInput.iloc[0] #grab the first row for the header
        dfInput = dfInput[1:] #take the data less the header row
        dfInput.columns = new_header #set the header row as the df header
        inputColumns = list(dfInput.columns)

        sheet_ranges = wb["Conversion file"] # This is the conversion file sheet of the primary deal data
        dfOutput = pd.DataFrame(sheet_ranges.values).head(2)
        new_header = dfOutput.iloc[0] #grab the first row for the header
        dfOutput = dfOutput[1:] #take the data less the header row
        dfOutput.columns = new_header #set the header row as the df header


        dfOutputColumns = list(dfOutput.columns)
        FunctionList = []
        ColumnfunctionList = []
        for index, row in dfOutput.iterrows():
            if index == 1:
                colIndex = 0
                for cell in row:
                   
                    functionObject = dict(self.createFunctionObject(cell))
                    # if not functionObject["isConstant"]:
                    #     functiondef = functionObject["function"]
                    #     functionObject.pop('function', None)
                    #     FunctionList.append({"functionId":functionid,"functiondef":functiondef})
                    ColumnfunctionList.append({"index":colIndex,"functionFormula": cell,"functionMetaData":functionObject})
                    colIndex=colIndex+1
        dyn_resource = self.getDynamo()
        try:
            table = dyn_resource.Table("psl-db-data-mapper-lab")
            # print("Size")
            # print(sys.getsizeof(FunctionList))

            table.put_item(
                    Item={
                        'Id': str(inputId),
                        'EngineUsed': "Excel",
                        'poolId': self.poolId,
                        'dealId':self.dealId,
                        'inputColumns':inputColumns,
                        'outputColumns': dfOutputColumns,
                        'columnFunctionData': ColumnfunctionList})



            # table = dyn_resource.Table("psl-db-function-definitions-lab")

            # for function in FunctionList:
            #     table.put_item(
            #     Item={
            #         'FunctionId': str(function["functionId"]),
            #         'functiondef': function["functiondef"],
            #         })

    
            # print(str(inputId))
            return pd.read_excel(self.fileData,sheet_name="Conversion file")
        except Exception as err:
            print(err)         


    def loadFunctionDict(self,data):
        # dyn_resource = self.getDynamo()
        # table = dyn_resource.Table("psl-db-function-definitions-lab")
        funcDict =dict()
        for columnDef in data["columnFunctionData"]:
        
            if not columnDef["functionMetaData"]["isConstant"]:
                # response = table.get_item(
                #     Key={'FunctionId': columnDef["functionid"]  }
                # )
                func = formulas.Parser().ast(columnDef["functionFormula"])[1].compile() 
                funcDict[columnDef["index"]]=func
        return funcDict



    def getVlookup(self,start,end):
        wb = load_workbook(BytesIO(self.fileData), 
                    read_only=True)

        ws = wb[wb.sheetnames[0]]

    # Read the cell values into a list of lists
        data_rows = []
        for row in ws[start:end]:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)

    # Transform into dataframe
        df = pd.DataFrame(data_rows)
        return df

    def is_number(self,str):
        ISNUM_REGEXP = re.compile(r'^[-+]?([0-9]+|[0-9]*\.[0-9]+)([eE][-+]?[0-9]+)?[ij]?$')

    #change order if you have a lot of NaN or inf to parse
        if ISNUM_REGEXP.match(str) or str == "NaN" or str == "inf": 
            return True 
        else:
            return False

    def buildDataFromFormula(self,depth,columnInfo,dfInput,functionDict,rowVal,rowIndex,columnIndex): # This function is used to build the data from the formula 
        parameterTuple = tuple() # This is the parameter tuple of the formula 
        # print("for column index  {index}".format(index=columnIndex))
        vlookupList = []
        # if columnIndex==24:
        #     print("reached")
        for parameter in columnInfo[columnIndex]["functionMetaData"]["parameter"]:
            if parameter["isVlookup"]:
                vlookupList.append(dill.loads( bytes(parameter["vlookUpHash"])) )
 
                parameterTuple = parameterTuple + (vlookupList[len(vlookupList)-1],)
            elif parameter["isSelfRef"]:
                if int(parameter["parameterIndex"])<len(rowVal):
                    parameterVal = rowVal[int(parameter["parameterIndex"])-1]
                    if pd.isna(parameterVal) or len(str(parameterVal))==0:
                        parameterVal=''
                    parameterTuple = parameterTuple + (parameterVal,)
                else:
                    parameterVal = self.buildDataFromFormula(depth+1, columnInfo  ,dfInput,functionDict,rowVal,rowIndex,int(parameter["parameterIndex"])-1)
                    if pd.isna(parameterVal) or len(str(parameterVal))==0:
                        parameterVal=''
                    parameterTuple = parameterTuple + (parameterVal,)
            else:
                parameterVal = dfInput.iloc[rowIndex,int(parameter["parameterIndex"])-1]
                if pd.isna(parameterVal) or len(str(parameterVal))==0:
                        parameterVal=''
                parameterTuple = parameterTuple + (parameterVal,)
        func = functionDict[columnIndex]
        result = func(*(parameterTuple))
        return result

    def SurveillanceDealData(self): # This function is used to process the surveillance deal data 
        dfInput = pd.read_excel(self.fileData,sheet_name="Data")    # This is the input data of the surveillance deal data
        dyn_resource = self.getDynamo() # This is the dynamo resource 
        table = dyn_resource.Table("psl-db-data-mapper-lab") 
        response = table.scan()["Items"]# Get all the items from the table 
        masterContent = None # This is the master content of the surveillance deal data
        for value in response:
            if value["poolId"]==self.poolId and value["dealId"]==self.dealId: # Check if the poolId and dealId matches the input 
                masterContent=value # This is the master content of the surveillance deal data
                break # This breaks the loop
        
        outputDf = pd.DataFrame(columns=list(masterContent["outputColumns"])) # This is the output data of the surveillance deal data
        functionDict = self.loadFunctionDict(masterContent) # This is the function dictionary of the surveillance deal data 
        # print(functionDict)
        for i in range(len(dfInput)): # This loop is used to process the surveillance deal data
            computedRowVal = dict() # This is the computed row value of the surveillance deal data 
            if i == 173: # This condition is used to check if the index is 173 
                pass # This is a pass statement
            testValLen =  len(str(dfInput.iloc[i,0])) # This is the length of the test value 
            testVal = str(dfInput.iloc[i,0]) # This is the test value 
            if dfInput.iloc[i,0]==None or (str(dfInput.iloc[i,0]))=='nan' or len(str(dfInput.iloc[i,0])) == 0 : # This condition is used to check if the test value is None or nan or the length of the test value is 0
                return outputDf # This returns the output data of the surveillance deal data
            for columnInfoindex in range(len(masterContent["columnFunctionData"])): # This loop is used to get the column information index of the surveillance deal data
                if masterContent["columnFunctionData"][columnInfoindex]["functionMetaData"]["isConstant"]: # This condition is used to check if the column information index is a constant
                    computedRowVal[columnInfoindex]=masterContent["columnFunctionData"][columnInfoindex]["functionMetaData"]["ConstantVal"]# This is the computed row value of the surveillance deal data 
                else: 
                    computedRowVal[columnInfoindex] = self.buildDataFromFormula(0,masterContent["columnFunctionData"],dfInput,functionDict,computedRowVal,i,columnInfoindex) # This is the computed row value of the surveillance deal data 
            outputDf.loc[i] = list(collections.OrderedDict(sorted(computedRowVal.items())).values())
            # print("index {index}".format(index=i))
        return outputDf



import time
filepath = "830365088_830365087_04262018_Surve.xlsm"
poolId = filepath.split("_")[0]
dealId = filepath.split("_")[1]
start_time = time.time()
DealProcessor(filepath,poolId,dealId).SurveillanceDealData()
print("--- %s seconds ---" % (time.time() - start_time))
