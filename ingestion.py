import collections
import datetime
import json
from openpyxl import load_workbook
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
import formulas
import dill
import string
# read file
import boto3
import os
from pathlib import Path
import time
from decimal import Decimal
from pathlib import Path
import sys
import subprocess
import re
import warnings
warnings.filterwarnings("ignore")





class Ingestion:
       
    def __init__(self,file_name):
     self.file_name = file_name
     self.template_col_len = 0
     self.df_meta_data=self.process()
     self.template_df, self.input_data_frame = self.Find_Input_Output_Dfs()
     self.ExtractColumns=self.ExtractColumns()
     


    def ExtractColumns(self):
        column_names = []
        for item in self.input_data_frame['ColumnData']:
            column_names.append(item['ColumnName'])
        return column_names

    def col2num(self,col):
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num
    

    def find_row_with_highest_count(self,ws_title):
        index =1
        index_dict =dict()
        wb = load_workbook(filename=file_name, read_only=True,data_only=True)
        ws=wb[ws_title]
        for row in ws:
            string_count =sum(isinstance(cell.value,str) for cell in row )
            index_dict[index]=string_count
            index+=1
            if index >= ws.max_row*0.05:
               break
        max_key = max(index_dict, key=index_dict.get)
        return max_key
 
        


    # def find_row_with_highest_count(self,ws):    
    #     index = 1
    #     index_dict=dict()
    #     for row in ws:
    #         unique_colors = set(cell.fill.start_color.index
    #                                    for cell in row
    #                                    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color)
    #         color_count =len(unique_colors)
    #         val_count = sum(cell.value is not None for cell in row)
             
    #         index_dict[index] = (val_count,color_count)
    #         index+=1  
    #         if index >= ws.max_row*0.05:
    #             break
    #     max_key = max(index_dict, key= lambda k: min(index_dict[k]))
    #     return max_key
        
        

    def initialize_df_column_metadata(self,column_name,has_formula,col_index,col_value):
        column_dict = dict()
        column_dict["ColumnName"] = column_name
        column_dict["ColValue"] = col_value
        column_dict["IsFormula"] = has_formula
        column_dict["ColLetter"] = get_column_letter(col_index)
        return column_dict
    def initialize_df_metadata(self,ws):
        df_dict = dict()
        df_dict["ValidColor"] = set()
        df_dict["ColumnData"] = list()
        df_dict["FormulasCount"] = 0
        df_dict["ColCount"] = 0
        df_dict["SheetName"]=ws.title
        return df_dict
    def init_maxkey(self,ws,max_key):
        maxkey_dict=dict()
        maxkey_dict[ws.title]=max_key
        return maxkey_dict
    

    
    # def Find_Input_Output_Dfs(self,current_df):
  
    #     df_meta_data=self.process(current_df,ws)
    #     temp_df = pd.DataFrame(df_meta_data)
    #     #temp_df.to_json('data.json', orient='records', lines=True)
    #     template_df =  temp_df.sort_values([ 'FormulasCount'],ascending=False).iloc[0].to_dict()
    #     #template_df.to_json('test.json', orient='records', lines=True)
    #     temp_df.drop( df_meta_data.index(template_df),inplace=True)
    #     input_data_frame =  temp_df.sort_values(['ColCount', 'FormulasCount'],ascending=False).iloc[0].to_dict()
    #     # todo : remove columns from input_data_frame which have formulas and create intermediate dfs and append all other dfs from temp_df
        
    #     # print(f"find_row_with_highest_count: {end_time - start_time} seconds")

    #     return template_df,input_data_frame
   

    def Find_Input_Output_Dfs(self):
        start_time = time.time()
        items_to_remove = []
        temp_df = pd.DataFrame(self.df_meta_data)
        template_df =  temp_df.sort_values([ 'FormulasCount'],ascending=False).iloc[0].to_dict()
        self.template_col_len = len(temp_df["ColCount"])
        temp_df.drop( self.df_meta_data.index(template_df),inplace=True)

        input_data_frame =  temp_df.sort_values(['ColCount', 'FormulasCount'],ascending=False).iloc[0].to_dict()
        temp_df_metadata = self.df_meta_data
        input_df_sheetname=input_data_frame['SheetName']
        output_df_sheetname=template_df['SheetName']
        temp_df_metadata = [item for item in self.df_meta_data if item['SheetName'] in [input_df_sheetname, output_df_sheetname]]

        temp_df_metadata.remove(input_data_frame)
        temp_df_metadata.remove(template_df)

        for item in input_data_frame.get('ColumnData', [])[:]:  
            if item.get('IsFormula', False):
                items_to_remove.append(item)
                input_data_frame['ColumnData'].remove(item)
        input_data_frame['ColCount'] = max(0, input_data_frame.get('ColCount', 0) - len(items_to_remove)) 
        

        for items_remove in  items_to_remove:
            template_df["ColumnData"].append(items_remove)


        for inter_df in  temp_df_metadata:
            for item in inter_df['ColumnData']:
                    item['IsIntermediate'] = True

        for inter_df in  temp_df_metadata:
            
            template_df["ColumnData"].extend(inter_df["ColumnData"])

        for item in template_df["ColumnData"]:
            if 'IsIntermediate' not in item:
                item['IsIntermediate'] = False


        end_time = time.time()  
        print(f"find_row_with_highest_count: {end_time - start_time} seconds")
        return template_df,input_data_frame
    


    def valid_color_check(self, col_index, max_key, current_df, ws):
        cell_fill = ws[max_key][col_index].fill
        if cell_fill is not None and cell_fill.start_color is not None:
            color_index = str(cell_fill.start_color.index)
            if (color_index in current_df["ValidColor"]) or (len(current_df["ValidColor"]) == 0):
                if color_index != '00000000':
                    return True
        return False

    def process(self):
        start_time = time.time()
        df_meta_data = list()
        wb = load_workbook(filename=file_name, read_only=True)
        for sheet in wb.sheetnames:    
            col_index = 0
            max_key =self.find_row_with_highest_count(sheet)
            ws = wb[sheet]
            current_df = self.initialize_df_metadata(ws)
            max_key_dict=self.init_maxkey(ws,max_key)
            max_key_dict[ws.title]=max_key
        
            while col_index<ws.max_column:
                has_formula = False
                if ws[max_key][col_index].value == None:
                    if  len(current_df["ColumnData"]) == 0:  
                        col_index += 1      
                    else:
                        if self.valid_color_check(col_index,max_key,current_df,ws):
                            current_df["ValidColor"].add( str(ws[max_key][col_index].fill.start_color.index ))
                            if str(ws[max_key+1][col_index].value).startswith('='):
                                current_df["FormulasCount"] += 1
                                has_formula = True
                            current_df["ColumnData"].append(self.initialize_df_column_metadata(ws[max_key][col_index].value,has_formula,col_index+1,str(ws[max_key+1][col_index].value)))
                            
                            current_df["ColCount"] += 1
                            current_df["SheetName"]=ws.title
                           
                            
                            col_index += 1
                        else:
                            df_meta_data.append(current_df)
                            current_df = self.initialize_df_metadata(ws)
                            col_index += 1      
                else :
                    try:
                        current_df["ValidColor"].add( str(ws[max_key][col_index].fill.start_color.index ))
                    except:
                        pass
                    if str(ws[max_key+1][col_index].value).startswith('='):
                        current_df["FormulasCount"] += 1
                        has_formula = True
                    current_df["ColumnData"].append(self.initialize_df_column_metadata(ws[max_key][col_index].value,has_formula,col_index+1,str(ws[max_key+1][col_index].value)))
                    current_df["ColCount"] += 1
                    current_df["SheetName"]=ws.title
                    col_index += 1
            if len(current_df["ColumnData"]) != 0:
                df_meta_data.append(current_df)
            end_time = time.time()
            print(f"Process: {end_time - start_time} seconds")
        return df_meta_data
        
        

    
    def process_all_sheets_except_active(self, file_name):
        all_sheets_meta_data = []

        for sheet_name in wb.sheetnames:  
            ws = wb[sheet_name]
            current_df = self.initialize_df_metadata(ws)
            sheet_meta_data = self.process(current_df, ws)

            sheet_meta_data_converted = [
                {k: list(v) if isinstance(v, set) else v for k, v in item.items() if k != "ValidColor"}
                for item in sheet_meta_data
            ] if isinstance(sheet_meta_data, list) else sheet_meta_data

            all_sheets_meta_data.append({"SheetName": sheet_name, "MetaData": sheet_meta_data_converted})
        all_sheets_meta_data_json = json.dumps(all_sheets_meta_data, default=self.handle_date_obj)
        all_sheets_meta_data = json.loads(all_sheets_meta_data_json, parse_float=Decimal)
        return all_sheets_meta_data

    def handle_date_obj(self, o):
        if isinstance(o, datetime.date):
            return o.isoformat()   
    
    # def merg_dfs(self):
    #     additional_dfs = self.process_all_sheets_except_active(file_name)
    #     if isinstance(additional_dfs, list) and all(isinstance(df, pd.DataFrame) for df in additional_dfs):
    #         all_dfs = [self.df_meta_data] + additional_dfs
    #         merged_df = pd.concat(all_dfs, ignore_index=True)
    #         return merged_df


    
    def getDynamo(self):
        return boto3.resource(
        'dynamodb',
        aws_access_key_id     = 'ASIA4C44A42VPPEXLJCK',
        aws_secret_access_key = 'GoVvzg3s9oOQ36FhbTMpZzjmuwshyySh3pifQYoz',
        aws_session_token     = 'FwoGZXIvYXdzEFUaDM/mQFa1PAtbL4rhVSKTAtmPud5nTlYakjsAMbMeKqKkOyBFNzyrfgmy2oZtNR+vpeBjGLTzKSsXpFBa9w1jLP5serHZfsf1QgGMbfxGc42u2YyUX+eGJShz8YzLSNiYKTKrFqUWMK9LY1pAq8EXMIbqYIoANyJf/0VdvzCUDCKnTK8lb/YLHVwlKXmwYpQXXXKCbbbFCMyRJVhe7HRRGIPc8Bu86fIrWJRrUBoEvN048nNM1UCJ2nUo7ilMJlPSxmuU6378xC5HzCRL4OVqn16qnM/Qaac+0G/9/j4DhLJLgdOfRNdL5HN61fChMl1x+pNuRrGsAJf9uUiZpNY523BQ4symHEAv50uRugfqQBKRr0QUfgpN3qwlepSpha2xSN+/KLrynbUGMitWf0RjWVcRRNr0otWxxs1vGI/vkahEQrxDRnYZDg/Df2dvThEb5vX8Y0PA',
      region_name="us-east-1")
    

    def createFunctionObject(self,formula,sheet_name,column_name):
        ws=wb[self.template_df["SheetName"]]
        parameterIndex = 0
        to_check_sheets = [sheet for sheet in wb.sheetnames if sheet.lower()!=sheet_name.lower() ]
        isConstant = False
        vLookupHash = None
        if len(str(formula))==0 or str(formula)[0]!='=':
            isConstant = True
            ConstantVal = formula
            return {"isConstant": isConstant,"ConstantVal": ConstantVal}
        func = formulas.Parser().ast(formula)[1].compile()
        parameters = []
        for parameter in func.inputs:
            isSelfRef = False
            isVlookup = False
            vLookupHash = None
            if   str(parameter).split('!')[0].upper().replace("'","") in (sheet.upper() for sheet in to_check_sheets):
                isVlookup = True
            
                try:
                  start,end = str(parameter).split('!')[1].split(":")
                  start = start.replace('$', '')
                  end= end.replace('$', '')
                  print([start, end])
                except:
                    pass
                try:
                 vLookupHash = dill.dumps( self.getVlookup(str(parameter).split('!')[0].replace("'",""),start,end))
                except:
                    pass
            elif  self.check_if_parameter_in_input(str(parameter)) :   
                parameterIndex = parameter
            elif self.check_if_parameter_in_template(str(parameter)):
                isSelfRef = True
                parameterIndex = parameter
            else:
                # todo handle this exception
                parameterIndex = parameter
        
            if not isVlookup and parameterIndex!=0:
                parameter_col = ws.cell(self.find_row_with_highest_count(ws.title),self.col2num(parameterIndex)).value
            else:
                 parameter_col = ""
            if parameter_col == None:
                parameter_col=""
            parameters.append({"parameterIndex":parameterIndex,"vlookUpHash":vLookupHash,"isSelfRef":isSelfRef,"isVlookup":isVlookup,"ParameterColumnName": parameter_col })
        return {"parameter": parameters,"isConstant":isConstant}
    
    def getVlookup(self,sheet_name,start,end):
        indx = next(i for i,v in enumerate(wb.sheetnames) if v.lower() == str(sheet_name).lower())
        ws = wb.worksheets[indx] 
        start=self.col2num(start)
        end=self.col2num(end)
        data_rows = []
        if  start == end:
            for row in ws[start:end]:
                data_rows.append(row.value)
        else:
            for row in ws[start:end]:
                data_cols = []
                for cell in row:
                    data_cols.append(cell.value)
                data_rows.append(data_cols)
        df = pd.DataFrame(data_rows)
        return df
    
    # def getVlookup(self, sheet_name, start_col, end_col):
    #     indx = next(i for i, v in enumerate(wb.sheetnames) if v.lower() == str(sheet_name).lower())
    #     ws = wb.worksheets[indx]

    #     start_col_index = self.col2num(start_col)
    #     end_col_index = self.col2num(end_col)
    #     data_rows = []
    #     for row in ws.iter_rows(min_col=start_col_index, max_col=end_col_index, values_only=True):
    #         if start_col_index == end_col_index:
    #             data_rows.append([cell for cell in row if cell is not None])  
    #         else:
    #             data_rows.append([cell if cell is not None else "" for cell in row])  
    #     df = pd.DataFrame(data_rows)
    #     df.to_json('test.json', orient='records', lines=True)
    #     return df
    

    def check_if_parameter_in_input(self,index):
        if (self.col2num(index) >=self.col2num(self.input_data_frame['ColumnData'][0]["ColLetter"])) and  (self.col2num(index) <=self.col2num(self.input_data_frame['ColumnData'][-1]["ColLetter"])):
            return True
        return False

    def check_if_parameter_in_template(self,index):
        
        if (self.col2num(index) >=self.col2num(self.template_df['ColumnData'][0]["ColLetter"])) and  (self.col2num(index) <=self.col2num(self.template_df['ColumnData'][self.template_col_len-1]["ColLetter"])):
            return True
        return False
    

    def PrimaryDealData(self):
        inputId = str(datetime.datetime.now().strftime('%Y%m%d%H%M%S%f'))  # This is the input id of the primary deal dat
        functionList = []
        ColumnfunctionList = []
        colIndex = 0  # Initialize column index
        for col in self.template_df["ColumnData"]:
            functionObject = dict(self.createFunctionObject(col["ColValue"], self.template_df["SheetName"], col["ColumnName"]))
            columnData = {
                "index": col["ColLetter"],
                "functionFormula": col["ColValue"],
                "functionMetaData": functionObject,
                "ColumnName": col["ColumnName"]
            }
            # Check if "IsIntermediate" is True or False in the template_df, and append accordingly
            if col.get("IsIntermediate")==False:
                columnData["IsIntermediate"] = False
            else:
                columnData["IsIntermediate"] = True
            ColumnfunctionList.append(columnData)
            colIndex += 1
        dyn_resource = self.getDynamo()
        #try:
        table = dyn_resource.Table("psl-db-data-mapper-lab")
        print("Size")
        print(sys.getsizeof(functionList))
        #all_sheets_meta_data = self.process_all_sheets_except_active(file_name)
        # self.ExtractColumns = [Decimal(x) if isinstance(x, float) else x for x in self.ExtractColumns]

        self.input_data_frame = {k: v for k, v in self.input_data_frame.items() if k != 'ValidColor'}
        self.template_df = {k: v for k, v in self.template_df.items() if k != 'ValidColor'}
        if is_dealId_poolId:
            deal_id=dealId
        else:
            deal_id=Path(file_name).stem



        table.put_item(
        Item={
            
                'Id': str(inputId),
                'template_df': self.template_df,
                'input_data_frame': self.input_data_frame,
                'inputColumns':self.ExtractColumns,
                #'All_Sheets_Meta_Data': self.process_all_sheets_except_active(file_name),
                #'All_Sheets_Meta_Data': self.merg_dfs(),
                'EngineUsed': "Excel",
                # 'poolId': self.poolId,
                'dealId': deal_id,
                'columnFunctionData': ColumnfunctionList
                }
            )



    def Similarity_Score(self):
        file_input_columns = self.ExtractColumns
        table_name = "psl-db-data-mapper-lab"
        dynamodb = self.getDynamo()
        table = dynamodb.Table(table_name)
        response = table.scan()
        items = response['Items']
        surveillance_triggered = False
        
        for item in items:
            if 'inputColumns' not in item:
                continue
            db_input_columns = item['inputColumns']
            intersection = len(set(file_input_columns) & set(db_input_columns))
            union = len(set(file_input_columns) | set(db_input_columns))
            similarity_percentage = (intersection / union) * 100
            print(similarity_percentage)

            if similarity_percentage > 90:
                subprocess.run([sys.executable, "templateextraction_1.py"])
                surveillance_triggered = True
                break
        if not surveillance_triggered:
            self.PrimaryDealData()

    def check_dealid_poolid(self,dealId,poolId):  
        dyn_resource = self.getDynamo()
        table = dyn_resource.Table("psl-db-data-mapper-lab")
        response = table.get_item(
            Key={
                'dealId': dealId  
            }
        )
        if 'Item' in response:
            print('True')
            inputId = response.get('Item', None).get('Id', None)
            print(inputId)
            subprocess.run([sys.executable, "templateextraction_1.py"])
            #self.SurveillanceDealData(self, inputId)
        else:
            self.PrimaryDealData()



  


start_time = time.time()
file_name = r"Apollo Series 2023-1 from_Suncorp MILAN Template - HL edit .xlsm"
split_name = file_name.split("_")
if len(split_name) >= 2:
    poolId = split_name[0]
    dealId = split_name[1]
else:
    poolId='-'
    dealId='-'
start_time = time.time()
wb = load_workbook(filename=file_name, read_only=True) 

# highest_return = -1
# highest_return_sheet_name = ""

# for sheet_name in wb.sheetnames:
#     df = pd.read_excel(file_name, sheet_name=sheet_name)
#     sheet_return = df.notnull().any().sum()
#     if sheet_return > highest_return:
#         highest_return = sheet_return
#         highest_return_sheet_name = sheet_name
# ws=wb.active
# print(ws.title)

if len(poolId) == 9 and len(dealId) == 9:
    is_dealId_poolId =True
    Ingestion(file_name).check_dealid_poolid(dealId,poolId) 
else:
    is_dealId_poolId=False
    Ingestion(file_name).Similarity_Score()
print("Data Ingested Successfully")
end_time = time.time()
print(f"Whole_Process_Time: {end_time - start_time} seconds")


# wb = load_workbook(filename=file_name,read_only=True) 
# ws = wb.active

# Ingestion(file_name).PrimaryDealData()
# end_time = time.time()
# print("Data Ingested Successfully:" +str(end_time-start_time))


# # raw_poolcut_file =r"input-2.xlsx"
# # raw_poolcut_file =r"raw_file.xlsx"

# split_name = file_name.split("_")
# if len(split_name) >= 2:
#     poolId = split_name[0]
#     dealId = split_name[1]
# else:
#     poolId='-'
#     dealId='-'




#Ingestion.SurveillanceDealData("CUA Harvey Template",file_name)
#Ingestion(file_name).Find_Input_Output_Dfs(current_df)
#Ingestion(file_name).Similarity_Score()
#Ingestion(file_name).put_data()
#Ingestion(file_name).put_data()



    
