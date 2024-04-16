import os
import io
import csv
import time
import gzip
import uuid
from pathlib import Path
import pandas as pd
import numpy as np
import pymongo as pm
from pymongo import MongoClient
from datetime import datetime
from openpyxl import load_workbook
from pfapiutils.handlers.XmlFileProcessor import process_xml
from const import dbconst
try:
    from pandas.core.common import maybe_box_datetimelike
except ImportError:
    from pandas.core.common import _maybe_box_datetimelike as maybe_box_datetimelike
from conf import env

# ToDo: Move Database connections to a different file
config = env.get_config()
client = MongoClient(f'{config.MONGODB_HOST}:{config.MONGODB_PORT}/', maxPoolSize=50, connect=True)
db = client[config.MONGODB_DATABASE_LDS]


def handle_xml(item, source = '', source_file_path ='', file_attribute=None, task_id=None):
    csvItem = process_xml(item, db)
    handle_csv(csvItem, source, source_file_path)

"""
    Steps to handle CSV files:
    1) Identify columns
    2) Identify Matching Rule - If rule does not exist create a new rule 
    3) Load data into corresponding table as defined in rule
    4) Create entry into FilesCollectionTable in Transaction with Data
"""
def handle_csv(item, source = '', source_file_path ='', file_attribute=None, task_id=None, parent_file_metadata_id = None):
    fname = item.replace('\\', '\\\\')
    '''f = fname[(len(fname) - fname.rindex('\\')-2)*-1:]
    if f.startswith('.~lock.'):
        pass    '''
    #tic1 = time.perf_counter()
    print(f'processing file: {fname} for data import')
    
    if '.gz' in fname:
        with gzip.open(fname) as f:
            df = pd.read_csv(f, encoding='latin1', nrows = 10)
            fileencoding = 'latin1'
    else:
        try:            
            df = pd.read_csv(fname, encoding='utf-8', nrows = 10)
            fileencoding = 'utf-8'
        except UnicodeDecodeError:            
            df = pd.read_csv(fname, encoding='latin1', nrows = 10) # latin1'
            fileencoding = 'latin1'
        except:            
            #log error
            pass
    
    # The trick here is that column rows can be anywhere in top 10 rows and find_columns defines from
    # which row our actual data starts.
    colloc = find_columns(df)

    # Best Column extractor identifies the logic to promote best suitable column for rule matching or creation.
    colresult = best_columnextractor(df,colloc)

    # TODO: to change this logic to loop and identify all the matching rules as against only 1st one for now.
    
    # Logic below identifies matching rule and if it does not exists then it creates 1 
    if colloc > 0:
        col = colresult.values[0]
        othercol = colresult.values[1:]
    else:
        col = colresult
        othercol = []
    
    rule = find_rule(col.tolist())
    
    lookupcategory = ""

    if len(rule) == 0:
        rule,insid = add_rule(col,othercol,'csv','','')
        collection = rule['Collection']
        score = 1.0
        id = insid.inserted_id
        rulename = rule['RuleName']
    elif len(rule) >=1:
        collection = rule[0]['Collection']
        score = rule[0]['score']
        id = rule[0]['id']
        rulename = rule[0]['RuleName']
        lookupcategory = rule[0]['LookupCategory']
    else:
        pass
    
    if '.gz' in fname:
        with gzip.open(fname) as f:            
            df = pd.read_csv(f, encoding=fileencoding)
            
    else:
        df = pd.read_csv(fname, encoding=fileencoding)
    
    rowcount = len(df.index)
        
    #Save metadata details in the database
    filemetadataid = log_schemadeatils(fname, source, col, othercol, colloc, rowcount, rulename, id, score, collection, source_file_path, fileencoding, file_attribute, task_id, parent_file_metadata_id)
    
    #Match the lookup data if we have lookupcategory available to match
    if(lookupcategory != ""):
        df = match_lookup_data(df, lookupcategory)
    df = trim_data(df, colloc, filemetadataid,True)
    
    #Save actual data in database
    save_data(df, collection, filemetadataid)    
    
    return filemetadataid
    #TODO: Additional code to remove the file

'''
    Steps to handle for XLSX:
    1) Identify Sheets and their respective columns
    2) Identify matching rules for each sheet - If not found create 1
    3) Load data into corresponding table as defined in Rule for each sheet where data is to be loaded
    4) Create Entry into FilesCollectionTable in Transaction with Data for each sheet where data is to be loaded
'''

def handle_xlsx(item, source = '', source_file_path ='', file_attribute=None, task_id=None):                                       
    fname = item.replace('\\', '\\\\')
    f = fname[(len(fname) - fname.rindex('\\')-2)*-1:]
    if f.startswith('.~lock.'):
        pass    
    
    print("xlsx file in dir: ", fname)
    wb = load_workbook(fname, read_only=True, keep_links=False, data_only=True)
    sheets = wb.sheetnames
    
    number_of_sheets = len(sheets)
    
    parent_file_metadata_id = None
    schemaId = None

    if(number_of_sheets > 1):  #treat the file like a 'normal' CSV file if number_of_sheets = 1        
        parent_file_metadata_id = log_schemadeatils(fname, source, [], [], 0, 0, None, None, None, None, source_file_path, None, file_attribute, task_id)

    for sheet in sheets:
        csv_file_name = Path(fname).stem + "." + sheet + ".csv"    
        wb = load_workbook(fname, read_only=True, data_only=True, keep_links=False)
        ws = wb[sheet]
        with open(csv_file_name, "w", newline='') as out:
            writer = csv.writer(out)        
            for row in ws.values:
                writer.writerow(row)
        print('About to handle CSV - ' + csv_file_name) 

        try:
            task_id = str(uuid.uuid4())  #for this temporary CSV file, create a new UUID for 'tid' 
            schemaId = handle_csv(csv_file_name, source, source_file_path, file_attribute, task_id, parent_file_metadata_id)
            print('Handled CSV - ' + csv_file_name + ' - with schemaId - ' + str(schemaId))                        
        except Exception as e:            
            print('There was a problem processing the file - ' + csv_file_name + '. Problem was - ' + repr(e))
        
    if parent_file_metadata_id is None:
        return schemaId
    return parent_file_metadata_id
   
def read_combine_excelsheets(xls_path):
    """Read all sheets of an Excel workbook and return a single DataFrame"""
    print(f'Loading {xls_path} into pandas')
    xl = pd.ExcelFile(xls_path)
    df = pd.DataFrame()
    columns = None
    for idx, name in enumerate(xl.sheet_names):
        print(f'Reading sheet #{idx}: {name}')
        sheet = xl.parse(name)
        if idx == 0:
            # Save column names from the first sheet to match for append
            columns = sheet.columns
        sheet.columns = columns
        # Assume index of existing data frame when appended
        df = df.append(sheet, ignore_index=True)
    return df

#Logic to identify the columns from a sheet/file. It only looks for first 10 rows and returns the row until which it sees column values
# The current logic is based on assumption such that their will be atleast 1 numeric or decimal data in the row to mark it as data row
def find_columns(df):
    i = 0
    while i >=0 and i < 10:
        df1 = df.iloc[[i]]
        #print(df1)
        dft = df1.applymap(type)
        #print(dft)
        v = list(set(dft.values[0]))
        #print(v)
        x = np.array(v)
        
        if x.size == 1:
            dft =  df1.applymap(lambda np: np.isnumeric() or np.isdecimal()) 
            #print(dft)
            v = list(set(dft.values[0]))
            #print(v)
            if len(v) > 1:
                break
            i += 1
        else:
            break
    return i

# Logic to identify which column should be choosen as the primary column incase their are multiple columns to choose from.
# Current logic is based on to identify the row which has least number of charecters in the column names so that it can be accomodated efficiently in MongoDB
def best_columnextractor(df, colloc):    
    colarr = df.iloc[:colloc].values
    if colarr.size>0:        
        colresult =  pd.concat([pd.DataFrame(colarr),pd.DataFrame([df.columns.values])]).reset_index(drop = True)
        
        noofcolumns = len(df.columns.values)
        
        uniqueval =  colresult.nunique(axis = 1)
        
        if len(set(uniqueval)) != 1:
            
            i = 0
            while i <= len(uniqueval):               
                
                if uniqueval[i] != noofcolumns:                    
                    colresult.drop(i, inplace=True)
                                                                
                else:
                    i+=1
    else:
        #Sanitize column values so that there are no errors while inserting
        colresult = df.columns.str.strip().str.lower().str.replace(' ', '_').str.replace('(', '').str.replace(')', '').str.replace('.', '')
        
    #print(colresult)

    #colresult will hold potential columns array.
    #Identify shortest column among potential rows but doing len comparision
    if colarr.size>0:
        colresult.loc[:,'t'] = colresult.applymap(lambda np: len(np)).sum(axis=1)
        colresult.sort_values('t', ascending=True, inplace=True)
        colresult.drop(['t'], axis = 1)
        #print(colresult)
        #Sanitize column values so that there are no errors while inserting/creating the new collection 
        
        colresult = colresult.applymap(lambda x: x.strip().lower().replace(' ', '_').replace('(', '').replace(')', '').replace('.', '') if isinstance(x, str) else x)
        
    return colresult

# Logic to identify matching rule (if any)
def find_rule(col):
    rule = db[dbconst.RULES]
    pipeline = [
        {'$unwind': { 'path': '$Columns'} },
        {'$match': {'Columns':{ '$in': col}}},
        {'$group': {'_id': {'id':'$_id', 'Priority':'$Priority', 'Collection':'$Collection', 'RuleName':'$RuleName', 'MatchingThreshold': '$MatchingThreshold', 'LookupCategory' : '$LookupCategory'}, 'count': {'$sum': 1}}},
        {'$project': {'_id': 0, 'id': '$_id.id', 'count': 1, 'Priority': '$_id.Priority', 'Collection':'$_id.Collection', 'RuleName':'$_id.RuleName','MatchingThreshold': '$_id.MatchingThreshold', 'LookupCategory' : '$_id.LookupCategory', 'score': {'$divide': ['$count', len(col)]}}},
        {'$match': { '$expr': { '$gte': [ '$score' , '$MatchingThreshold' ] } } },
        {'$sort': {'score': -1}}
        ]
    cursor = rule.aggregate(pipeline)
    return list(cursor)

# Logic to add a new rule in the rules collection
def add_rule(cols, other_columns, file_type, table_name, othertabn ):
    rule = db[dbconst.RULES]
    ruledoc = {
    'RuleName': 'AutomatedRule::' + str(datetime.utcnow()),
    'Columns': cols.tolist(),
    'OtherPotentialColumns': other_columns.tolist() if len(other_columns) > 0 else [],
    'Priority': 100,
    'Collection': 'automated_coll_' + str(int(datetime.utcnow().strftime("%Y%m%d%H%M%S"))),
    'TargetColumnReMapping' : '',
    'RemoveColumnChars': '',
    'Type': 'Process',
    'CtBy': 'Process',
    'CtAt': datetime.utcnow().isoformat(),
    'v': 1,
    'MatchingThreshold':.8,
    'LookupCategory': ''
    }
    print(ruledoc)
    i = rule.insert_one(ruledoc)
    return ruledoc, i
 
# Logic to trim the data so that it saves the space while saving the data in MongoDB
def trim_data(df, column_index, metadata_id,reduce_data_space = True):
    df['_s']  = metadata_id
    df = df.iloc[column_index:]
    #explicitly delete all empty columns to reduce space
    if reduce_data_space:
        df = df.dropna(axis=1,how='all') #explicitly delete all empty columns to reduce space in mongo
        df = df.replace(np.nan, '', regex=True)
        # todo: All these ND values should come from rule
        df = df.replace(dict.fromkeys(['ND1','ND2','ND3','ND4','ND5','ND,1','ND,2','ND,3','ND,4','ND,5','ND,6','ND,7'], ''))
    return df

# Logic to save the raw data into respective collection
def save_data(df, mongodb_collection, metadata_id, chunk_size = 1010 ):
    my_list1 = df.to_dict(orient='records')
    my_list = [{k:v for k,v in d.items() if v != '' } for d in my_list1]
    ct = len(df.index)
    chunk_size = chunk_size if ct > chunk_size else ct
    l =  len(df)                  
    ran = range(l + 1)
    steps=ran[chunk_size::chunk_size]
    # Insert chunks of the dataframe
    i = 0
    for j in steps:
        try:
            db[mongodb_collection].insert_many(my_list[i:j]) # fill thr collection
        except pm.errors.BulkWriteError as bwe:
            print(bwe.details)
            # todo: Log error
        
        i = j
    pass

# Logic to save metadata details
def log_schemadeatils(file_name, source, columns, other_columns, column_rows, row_count, matching_rule, matching_rule_id, matching_score, collection, source_file_path, encoding, file_attribute=None, task_id=None, parent_file_metadata_id = None):
    filemetadata = db[dbconst.FILE_METADATA]

    if isinstance(columns, list) == False:
        columns = columns.tolist()

    if isinstance(other_columns, list) == False:
        other_columns = other_columns.tolist()
    
    filemetadatadoc = {
        'filename': file_name,
        'source' : source,
        'sourcepath': source_file_path,
        'columns' : columns if columns is not None else [],
        'othercolumns' : other_columns if len(other_columns) > 0 else [],
        'columnrows' : column_rows,
        'rowcount': row_count,
        'matchingrule': matching_rule,
        'matchingruleid': matching_rule_id,
        'matchingscore': matching_score,
        'collection':collection,
        'createdat': datetime.utcnow().isoformat(),
        'createdby': 'System',
        'encoding': encoding,
        'fileattr':file_attribute,
        '_tid':task_id,
        'isactive':True,
        'parentfilemetadataid':parent_file_metadata_id
    }
    # check if file metadata already exists and associated with tid
    if task_id:
        fmetadata = filemetadata.find_one({"_tid":task_id})
        if not fmetadata:
            id = filemetadata.insert_one(filemetadatadoc).inserted_id
            print(f'file metadata created for file {file_name} with id {id}')
        else :
            print(f'File metadata with id: {fmetadata["_id"]} already exists ')
            print('Deleting existing LBL records (if any)')
            db[fmetadata['collection']].delete_many({'_s':{'$eq':fmetadata["_id"]}})
            id=fmetadata["_id"]
    else:
        id = filemetadata.insert_one(filemetadatadoc).inserted_id
        print(f'file metadata created for file {file_name} with id {id}')
    return id

# Logic to get lookupdata
def get_lookupdata(category):
    lkp = db.lookup
    cur = lkp.find({'category' : category })
    return list(cur)

# Logic to get matched lookup data
def match_lookup_data(df, lookup_category):
    data = get_lookupdata(lookup_category)
    lkpdf = pd.DataFrame(data)
    # Get unique column names which needs to be changed  
    columnNames = lkpdf['cn'].unique().tolist()
    for colName in columnNames:
        lkdf = lkpdf[lkpdf['cn'] == colName]
        lkdf = lkdf[['key','value']]
        newcolname = '_' + colName
        #rename the column names so that it prefixes _ with the columnName
        lkdf.rename(columns={"key": colName, "value": newcolname}, inplace = True)
        if(is_numeric_or_decimal(lkdf,True,colName)):
            lkdf[colName]=lkdf[colName].astype(int)
        elif(is_numeric_or_decimal(lkdf,False,colName)):
            lkdf[colName]=lkdf[colName].astype(float)
        # Other datatypes can also be added in the same way.
        df=add_string_datatype(newcolname,lkdf,colName,lkpdf)
    return df
def add_string_datatype(newcolname,lkdf,colName,lkpdf):
    error = False
    if newcolname in df.columns:
        filter_col = [col for col in df if not col.startswith(newcolname)]
        df = df[filter_col]
    #Incase of NDs it can be problamatic to merge if there is change in datatype 
    if colName in df.columns:
        try:
            df = pd.merge(df, lkdf, on=colName, suffixes = ('', 'Name'), how='left')
        except:
            lkdf[colName]=lkdf[colName].astype(str)
            error = True
        if(error):
            try:
                # If it fails then again try with string datatype
                df = pd.merge(df, lkdf, on=colName, suffixes = ('', 'Name'), how='left')
            except:
                #Log Error
                print('Error in matchlookupdata: ' + lkpdf['category'] + ' cn ' + colName)
                pass
    return df



def is_numeric_or_decimal(lkdf, numeric, column_name):
    if(numeric):
        dft =  lkdf.applymap(lambda np: np.isnumeric())
    else:
        dft =  lkdf.applymap(lambda np: np.isdecimal())
    v = dft[column_name].unique().tolist()
    if len(v) > 1:
        return False
    else:
        return True

# Entry point to the program and has different file handlers
def doprocess(filename, source, source_file_path, file_attribute=None, task_id=None):
    if '.csv' in filename:
        return handle_csv(filename, source, source_file_path, file_attribute, task_id)        
    if '.xlsx' in filename:
        return handle_xlsx(filename, source, source_file_path, file_attribute, task_id)
    if '.xml' in filename:
        return handle_xml(filename, source, source_file_path, file_attribute, task_id)
        #pass

# Uncomment to manually verify files from a particular folder    
'''
SEPERATOR = '\\'
location = 'C:\\Tools\\edw'
csvfiles_in_dir = []
xlsxfiles_in_dir = []

# r=>root, d=>directories, f=>files
for r, d, f in os.walk(location):
   for item in f:
      if '.csv' in item:
         csvfiles_in_dir.append(os.path.join(r, item))
      elif '.xlsx' in item:
         xlsxfiles_in_dir.append(os.path.join(r, item))

for csvfile in csvfiles_in_dir:
    handle_csv(csvfile)

for xlsxfile in xlsxfiles_in_dir:
    pass
'''